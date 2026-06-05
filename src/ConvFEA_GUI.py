"""
ConvFEA — Interface graphique pour études de convergence par éléments finis
Fichier unique, compatible PyInstaller.

Dépendances : tkinter (stdlib), matplotlib, pandas, openpyxl, reportlab, ansys-mechanical-core
Génération .exe : pyinstaller --onefile --windowed --icon=icon.ico ConvFEA_GUI.py
"""

# ─────────────────────────────────────────────────────────────────────────────
#  IMPORTS
# ─────────────────────────────────────────────────────────────────────────────
import sys
import os
import ctypes
import time
import threading
import queue
import textwrap
import json
import ansys.mechanical.core

import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import numpy as np
from scipy.interpolate import PchipInterpolator
import mplcursors

import pandas as pd
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
import matplotlib.gridspec as gridspec

from openpyxl import Workbook

# ─────────────────────────────────────────────────────────────────────────────
#  SCRIPT MECHANICAL EMBARQUÉ (mechanical_script.py – ne pas modifier)
# ─────────────────────────────────────────────────────────────────────────────
def get_resource_path(relative_path):
    """Obtient le chemin absolu vers la ressource, compatible dev et PyInstaller."""
    try:
        # PyInstaller crée un dossier temporaire et stocke le chemin dans _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def load_mechanical_script():
    script_path = get_resource_path("mechanical_script.py")
    with open(script_path, "r", encoding="utf-8") as f:
        return f.read()

MECHANICAL_SCRIPT = load_mechanical_script()

# --- Activation du High-DPI sous Windows ---
try:
    ctypes.windll.shcore.SetProcessDpiAwareness(1)
except Exception:
    pass

# --- Configuration globale du thème ---
ctk.set_appearance_mode("Dark")  # Force le mode sombre (ou "System", "Light")
ctk.set_default_color_theme("blue")  # Thème d'accentuation par défaut

# ─────────────────────────────────────────────────────────────────────────────
#  PALETTE & CONSTANTES DE STYLE
# ─────────────────────────────────────────────────────────────────────────────
DARK_BG     = "#0D1117"
PANEL_BG    = "#161B22"
CARD_BG     = "#1C2333"
BORDER      = "#30363D"
ACCENT      = "#58A6FF"
ACCENT2     = "#3FB950"
WARN        = "#F78166"
TEXT_PRI    = "#E6EDF3"
TEXT_SEC    = "#8B949E"
TEXT_DIM    = "#484F58"

FONT_MONO   = ("Consolas", 9)
FONT_LABEL  = ("Segoe UI", 9)
FONT_TITLE  = ("Segoe UI Semibold", 10)
FONT_H1     = ("Segoe UI Semibold", 12)
FONT_BADGE  = ("Segoe UI", 8)

# Couleurs des courbes matplotlib
CURVE_COLORS = [
    "#58A6FF", "#3FB950", "#F78166", "#D2A8FF",
    "#FFA657", "#79C0FF", "#56D364", "#FF7B72",
]

# Noms de colonnes du DataFrame
COL_NAMES = [
    "Taille maillage (mm)",
    "Contrainte max Von Mises (MPa)",
    "Contrainte moy Von Mises (MPa)",
    "Déplacement maximal (mm)",
    "Énergie de déformation (mJ)",
    "Min Qualité élément",
    "Moy Qualité élément",
    "Max Aspect Ratio",
    "Moy Aspect Ratio",
    "Min Jacobian Ratio",
    "Moy Jacobian Ratio",
    "Max Skewness",
    "Moy Skewness",
    "Nombre d'éléments",
]

# Définition des graphiques à tracer : (titre, colonnes_y, ylabel)
PLOT_CONFIGS = [
    ("Contrainte max Von Mises",
     ["Contrainte max Von Mises (MPa)"], "MPa"),
    ("Contrainte moy Von Mises",
     ["Contrainte moy Von Mises (MPa)"], "MPa"),
    ("Déplacement maximal",
     ["Déplacement maximal (mm)"], "mm"),
    ("Énergie de déformation",
     ["Énergie de déformation (mJ)"], "mJ"),
    ("Qualité des éléments",
     ["Min Qualité élément", "Moy Qualité élément"], "—"),
    ("Aspect Ratio",
     ["Max Aspect Ratio", "Moy Aspect Ratio"], "—"),
    ("Jacobian Ratio",
     ["Min Jacobian Ratio", "Moy Jacobian Ratio"], "—"),
    ("Skewness",
     ["Max Skewness", "Moy Skewness"], "—"),
]


# ─────────────────────────────────────────────────────────────────────────────
#  MOTEUR DE CALCUL (thread séparé)
# ─────────────────────────────────────────────────────────────────────────────
def run_study(mechdat_path: str, sizes: list, sim_name: str, show_gui: bool,
              log_q: queue.Queue, prog_q: queue.Queue, result_q: queue.Queue):
    """
    Exécuté dans un thread worker.
    Communique via trois queues :
      log_q    → messages texte pour la console
      prog_q   → (current, total) pour la barre de progression
      result_q → DataFrame final ou Exception
    """
    def log(msg): log_q.put(msg)

    try:
        from ansys.mechanical.core import launch_mechanical
    except ImportError:
        result_q.put(ImportError(
            "ansys-mechanical-core introuvable.\n"
            "Installez-le : pip install ansys-mechanical-core"
        ))
        return

    log("▶ Lancement d'Ansys Mechanical…")
    try:
        mechanical = launch_mechanical(batch=not show_gui)
    except Exception as e:
        result_q.put(RuntimeError(f"Impossible de lancer Mechanical : {e}"))
        return

    log(f"✔ Mechanical démarré (Mode Visuel: {show_gui}).")
    log(f"▶ Ouverture du modèle : {mechdat_path}")
    try:
        mechanical.run_python_script(
            f'ExtAPI.DataModel.Project.Open(r"{mechdat_path}")'
        )
    except Exception as e:
        result_q.put(RuntimeError(f"Erreur ouverture modèle : {e}"))
        mechanical.exit()
        return

    log("✔ Modèle chargé.")

    rows = []
    total = len(sizes)

    for idx, size in enumerate(sizes, 1):
        log(f"── Taille {size} mm ({idx}/{total})…")
        script = f"SIZE = {size}\n" + MECHANICAL_SCRIPT
        try:
            res_string = mechanical.run_python_script(script)
            # Lecture et décodage sécurisé du JSON d'Ansys
            data = json.loads(res_string.strip())
            
            row = [
                size, data["stress_max"], data["stress_avg"], data["disp_max"], data["energy_max"],
                data["eq_min"], data["eq_avg"], data["ar_max"], data["ar_avg"], data["jr_min"],
                data["jr_avg"], data["sk_max"], data["sk_avg"], data["elements"]
            ]
            rows.append(row)
            log(f"   ✔ Eléments: {data['elements']} | σmax={data['stress_max']:.2f} MPa | umax={data['disp_max']:.3f} mm | E={data['energy_max']:.3f} mJ")
        except Exception as e:
            log(f"   ✗ Erreur taille {size} mm : {e}")

        prog_q.put((idx, total))

    log("▶ Fermeture de Mechanical…")
    try:
        mechanical.exit()
    except Exception:
        pass
    log("✔ Mechanical fermé.")

    if not rows:
        result_q.put(RuntimeError("Aucun résultat collecté."))
        return

    df = pd.DataFrame(rows, columns=COL_NAMES)
    result_q.put(df)


# ─────────────────────────────────────────────────────────────────────────────
#  ANALYSE DE CONVERGENCE
# ─────────────────────────────────────────────────────────────────────────────
def compute_convergence(df: pd.DataFrame) -> pd.DataFrame:
    """Calcule les erreurs relatives (%) entre raffinements successifs."""
    dfc = df.copy()
    numeric_cols = COL_NAMES[1:]  # exclure la colonne taille
    for col in numeric_cols:
        dfc[f"Δ {col} (%)"] = dfc[col].pct_change().abs() * 100
    return dfc


def build_diagnostics(df: pd.DataFrame) -> list:
    """Génère des messages de diagnostic automatiques."""
    msgs = []
    size_col = "Taille maillage (mm)"

    def last_rel_err(col):
        vals = df[col].dropna().values
        if len(vals) < 2:
            return None
        ref = vals[-1]
        if ref == 0:
            return None
        return abs(vals[-1] - vals[-2]) / abs(ref) * 100

    # Convergence énergie
    err_e = last_rel_err("Énergie de déformation (mJ)")
    if err_e is not None:
        if err_e < 1:
            msgs.append(("✔", ACCENT2, "Convergence excellente de l'énergie de déformation (< 1 %)."))
        elif err_e < 5:
            msgs.append(("✔", ACCENT2, f"Convergence satisfaisante de l'énergie de déformation ({err_e:.1f} %)."))
        else:
            msgs.append(("⚠", WARN, f"L'énergie de déformation n'a pas convergé ({err_e:.1f} % entre les deux derniers raffinements)."))

    # Convergence contrainte moy
    err_avg = last_rel_err("Contrainte moy Von Mises (MPa)")
    if err_avg is not None:
        if err_avg < 2:
            msgs.append(("✔", ACCENT2, f"Convergence acceptable des contraintes moyennes ({err_avg:.1f} %)."))
        elif err_avg < 8:
            msgs.append(("⚠", "#FFA657", f"Contraintes moyennes en cours de convergence ({err_avg:.1f} %)."))
        else:
            msgs.append(("✗", WARN, f"Contraintes moyennes non convergées ({err_avg:.1f} %)."))

    # Contrainte max — singularité ?
    err_max = last_rel_err("Contrainte max Von Mises (MPa)")
    if err_max is not None:
        if err_max > 10:
            msgs.append(("⚠", WARN,
                "Les contraintes maximales continuent d'évoluer fortement — probable singularité de contrainte. "
                "Vérifiez la géométrie ou la zone de scoping."))
        elif err_max < 3:
            msgs.append(("✔", ACCENT2, f"Contraintes maximales stables ({err_max:.1f} %)."))

    # Qualité maillage
    aq_vals = df["Moy Qualité élément"].dropna().values
    if len(aq_vals):
        last_q = aq_vals[-1]
        if last_q >= 0.8:
            msgs.append(("✔", ACCENT2, f"Qualité des éléments globalement excellente (moy = {last_q:.3f})."))
        elif last_q >= 0.6:
            msgs.append(("✔", ACCENT2, f"Qualité des éléments acceptable (moy = {last_q:.3f})."))
        else:
            msgs.append(("✗", WARN, f"Qualité des éléments médiocre (moy = {last_q:.3f}). Raffinez ou retravaillez la géométrie."))

    # Jacobian
    jac_vals = df["Moy Jacobian Ratio"].dropna().values
    if len(jac_vals):
        last_j = jac_vals[-1]
        if last_j >= 0.5:
            msgs.append(("✔", ACCENT2, f"Jacobian Ratio dans une plage acceptable (moy = {last_j:.3f})."))
        else:
            msgs.append(("⚠", WARN, f"Jacobian Ratio potentiellement problématique (moy = {last_j:.3f})."))

    # Skewness
    sk_vals = df["Max Skewness"].dropna().values
    if len(sk_vals):
        last_sk = sk_vals[-1]
        if last_sk < 0.6:
            msgs.append(("✔", ACCENT2, f"Skewness maximal acceptable ({last_sk:.3f} < 0,6)."))
        else:
            msgs.append(("⚠", WARN, f"Skewness maximal élevé ({last_sk:.3f}) — risque de dégradation de précision."))

    # Aspect Ratio
    ar_vals = df["Max Aspect Ratio"].dropna().values
    if len(ar_vals):
        last_ar = ar_vals[-1]
        if last_ar > 20:
            msgs.append(("⚠", WARN, f"Aspect Ratio maximal très élevé ({last_ar:.1f}) — éléments aplatis."))
        else:
            msgs.append(("✔", ACCENT2, f"Aspect Ratio dans des limites raisonnables ({last_ar:.1f})."))

    if not msgs:
        msgs.append(("—", TEXT_SEC, "Pas assez de données pour un diagnostic complet."))

    return msgs


# ─────────────────────────────────────────────────────────────────────────────
#  UTILITAIRES TKINTER
# ─────────────────────────────────────────────────────────────────────────────
def styled_frame(parent, fg_color="#161B22", corner_radius=8, **kw):
    return ctk.CTkFrame(parent, fg_color=fg_color, corner_radius=corner_radius, **kw)


def h_sep(parent, bg=BORDER):
    return tk.Frame(parent, bg=bg, height=1)


def label(parent, text, font_size=12, text_color="#E6EDF3", bold=False, **kw):
    kw.pop("font", None)
    weight = "bold" if bold else "normal"
    font = ctk.CTkFont(family="Segoe UI", size=font_size, weight=weight)
    return ctk.CTkLabel(parent, text=text, font=font, text_color=text_color, **kw)

def entry(parent, textvariable=None, width=200, **kw):
    return ctk.CTkEntry(parent, 
                        textvariable=textvariable,
                        width=width,
                        fg_color="#1C2333",
                        border_color="#30363D",
                        text_color="#E6EDF3",
                        corner_radius=6,
                        **kw)


def accent_btn(parent, text, command, width=None, color="#58A6FF", hover_color="#79C0FF", **kw):
    cfg = dict(
        text=text,
        command=command,
        fg_color=color,
        hover_color=hover_color,
        text_color="#0D1117",
        font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
        corner_radius=6
        # On a retiré le height=36 codé en dur ici
    )
    if width:
        cfg["width"] = width
        
    # On définit 36 comme hauteur par défaut SEULEMENT si tu n'as pas précisé de height
    kw.setdefault("height", 36)
    
    return ctk.CTkButton(parent, **cfg, **kw)


def ghost_btn(parent, text, command, **kw):
    return ctk.CTkButton(parent, 
                         text=text, 
                         command=command,
                         fg_color="transparent", 
                         hover_color="#30363D",
                         text_color="#8B949E",
                         border_width=1,
                         border_color="#30363D",
                         corner_radius=6,
                         **kw)


# ─────────────────────────────────────────────────────────────────────────────
#  APPLICATION PRINCIPALE
# ─────────────────────────────────────────────────────────────────────────────
class ConvFEA(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("ConvFEA — Étude de convergence FEA")
        self.geometry("1420x860")
        self.minsize(1100, 700)

        # État interne
        self.df: pd.DataFrame | None = None
        self.history = {}
        self.current_loaded_name = None
        self.running = False
        self._log_q: queue.Queue = queue.Queue()
        self._prog_q: queue.Queue = queue.Queue()
        self._res_q: queue.Queue = queue.Queue()

        # Variables Tk
        self.var_path    = ctk.StringVar(value="")
        self.var_simname = ctk.StringVar(value="Simulation_1")
        self.var_mode    = ctk.StringVar(value="Plage")  # "range" | "list"
        self.var_min     = ctk.StringVar(value="5")
        self.var_max     = ctk.StringVar(value="25")
        self.var_step    = ctk.StringVar(value="5")
        self.var_list    = ctk.StringVar(value="5;10;15;20;25")
        self.var_prog    = ctk.DoubleVar(value=0)
        self.var_show_gui = ctk.BooleanVar(value=False)

        for var in [self.var_min, self.var_max, self.var_step, self.var_list]:
            var.trace_add("write", lambda *_: self._refresh_preview())

        self._build_ui()
        self._poll()  # démarrer la boucle de polling

    # ── Construction de l'interface ──────────────────────────────────────────
    def _build_ui(self):
        self._build_topbar()
        content = styled_frame(self, fg_color=DARK_BG)
        content.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        content.columnconfigure(1, weight=1)
        content.rowconfigure(0, weight=1)

        self._build_left_panel(content)
        self._build_center_area(content)
        self._build_right_panel(content)

    def _build_topbar(self):
        bar = styled_frame(self, fg_color=PANEL_BG)
        bar.pack(fill="x", pady=(0, 2))
        tk.Frame(bar, bg=ACCENT, width=4).pack(side="left", fill="y")
        label(bar, "ConvFEA", font=("Segoe UI Semibold", 14), text_color=ACCENT,
              fg_color=PANEL_BG).pack(side="left", padx=(14, 4), pady=10)
        label(bar, "Automatisation des études de convergence par éléments finis",
              text_color=TEXT_SEC, fg_color=PANEL_BG, font=FONT_LABEL).pack(side="left")
        label(bar, "v1.0", text_color=TEXT_DIM, fg_color=PANEL_BG,
              font=FONT_BADGE).pack(side="right", padx=16)

    # ── Panneau gauche ───────────────────────────────────────────────────────
    def _build_left_panel(self, parent):
        panel = ctk.CTkScrollableFrame(
            parent, 
            fg_color="#161B22",   # Votre couleur PANEL_BG
            width=260,            # Largeur fixe de la barre latérale
            corner_radius=8
        )
        panel.grid(row=0, column=0, sticky="nsew", padx=(0, 6))

        def section(title):
            tk.Frame(panel, bg=BORDER, height=1).pack(fill="x", padx=10, pady=(12, 6))
            label(panel, title, font=FONT_TITLE, text_color=ACCENT, fg_color=PANEL_BG).pack(
                anchor="w", padx=14, pady=(0, 6))

        # Titre
        label(panel, "Paramètres", font=FONT_H1, text_color=TEXT_PRI, fg_color=PANEL_BG).pack(
            anchor="w", padx=14, pady=(14, 2))

        # ── Modèle
        section("① Modèle mécanique")
        label(panel, "Fichier .mechdat", text_color=TEXT_SEC, fg_color=PANEL_BG).pack(anchor="w", padx=14)
        row = styled_frame(panel, fg_color=PANEL_BG)
        row.pack(fill="x", padx=14, pady=(4, 0))
        self._path_entry = entry(row, textvariable=self.var_path, width=18)
        self._path_entry.pack(side="left", fill="x", expand=True)
        ghost_btn(row, "…", self._browse, width=28).pack(side="left", padx=(4, 0))

        self._gui_switch = ctk.CTkSwitch(panel, text="Afficher l'interface Ansys", variable=self.var_show_gui, progress_color=ACCENT, text_color=TEXT_PRI, font=ctk.CTkFont("Segoe UI", 11))
        self._gui_switch.pack(anchor="w", padx=14, pady=(8, 0))

        # ── Nom simulation
        section("② Simulation")
        label(panel, "Nom de la simulation", text_color=TEXT_SEC, fg_color=PANEL_BG).pack(anchor="w", padx=14)
        entry(panel, textvariable=self.var_simname).pack(fill="x", padx=14, pady=(4, 0))

        # ── Tailles de maillage
        section("③ Tailles de maillage")
        # On définit la variable par défaut (attention, elle doit correspondre aux textes du bouton)
        self.var_mode.set("Plage") 

        # Création du bouton segmenté moderne
        self._mode_switch = ctk.CTkSegmentedButton(
            panel,
            values=["Plage", "Liste custom"],
            variable=self.var_mode,
            command=self._toggle_mesh_mode,
            selected_color="#58A6FF",          # Couleur ACCENT quand sélectionné
            selected_hover_color="#79C0FF",    # Couleur au survol
            unselected_color="#1C2333",        # Couleur de fond par défaut (CARD_BG)
            unselected_hover_color="#30363D",  # Survol quand non sélectionné
            text_color="#E6EDF3",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold")
        )
        self._mode_switch.pack(fill="x", padx=14, pady=(0, 12))

        # --- Conteneur fixe ---
        self._mesh_inputs_container = styled_frame(panel, fg_color=PANEL_BG)
        self._mesh_inputs_container.pack(fill="x", padx=14)

        # Sous-frame plage
        self._range_frame = styled_frame(self._mesh_inputs_container, fg_color=PANEL_BG)
        self._range_frame.pack(fill="x", padx=14)
        for lbl, var in [("Min (mm) :", self.var_min),
                          ("Max (mm) :", self.var_max),
                          ("Pas (mm) :", self.var_step)]:
            r = styled_frame(self._range_frame, fg_color=PANEL_BG)
            r.pack(fill="x", pady=2)
            label(r, lbl, text_color=TEXT_SEC, fg_color=PANEL_BG, width=10, anchor="w").pack(side="left")
            entry(r, textvariable=var, width=50).pack(side="left", padx=(15, 0))

        # Sous-frame liste
        self._list_frame = styled_frame(self._mesh_inputs_container, fg_color=PANEL_BG)
        label(self._list_frame, "Tailles séparées par « ; »",
              text_color=TEXT_SEC, fg_color=PANEL_BG).pack(anchor="w")
        entry(self._list_frame, textvariable=self.var_list, width=26).pack(
            fill="x", pady=(4, 0))

        # ── Aperçu des tailles
        section("④ Aperçu")
        self._sizes_preview = tk.Text(panel, height=4, bg=CARD_BG, fg=TEXT_SEC,
                                       font=FONT_MONO, relief="flat", wrap="word",
                                       state="disabled", bd=0, padx=6, pady=4)
        self._sizes_preview.pack(fill="x", padx=14, pady=(0, 4))
        ghost_btn(panel, "↻ Recalculer l'aperçu", self._refresh_preview).pack(
            anchor="w", padx=14, pady=(0, 4))
        self._toggle_mesh_mode()
        
        self._refresh_preview()

        # ── Lancement
        section("")
        self._run_btn = accent_btn(panel, "▶  Lancer l'étude de convergence",
                                    self._start_study, color=ACCENT2)
        self._run_btn.pack(fill="x", padx=14, pady=(0, 6))

        # Barre de progression
        style = ttk.Style()
        style.theme_use("default")
        style.configure("ConvFEA.Horizontal.TProgressbar",
                        troughcolor=CARD_BG, background=ACCENT2,
                        thickness=6, borderwidth=0)
        self._prog_bar = ttk.Progressbar(panel, variable=self.var_prog,
                                          style="ConvFEA.Horizontal.TProgressbar",
                                          mode="determinate", maximum=100)
        self._prog_bar.pack(fill="x", padx=14, pady=(0, 4))
        self._prog_label = label(panel, "En attente…", text_color=TEXT_DIM, fg_color=PANEL_BG,
                                  font=FONT_BADGE)
        self._prog_label.configure(anchor="w")
        self._prog_label.pack(anchor="w", padx=14)

        ghost_btn(panel, "💾 Enregistrer la simulation", self._save_current_to_history).pack(fill="x", padx=14, pady=(8, 4))

        # ── Export
        section("⑤ Export")
        for txt, cmd in [("Export CSV",   self._export_csv),
                          ("Export Excel", self._export_excel),
                          ("Export PDF",   self._export_pdf),
                          ("Export PNG",   self._export_png)]:
            ghost_btn(panel, txt, cmd).pack(fill="x", padx=14, pady=2)

    # ── Zone centrale ────────────────────────────────────────────────────────
    def _build_center_area(self, parent):
        center = styled_frame(parent, fg_color=DARK_BG)
        center.grid(row=0, column=1, sticky="nsew")
        center.rowconfigure(0, weight=1)
        center.rowconfigure(1, weight=0)
        center.columnconfigure(0, weight=1)

        # Notebook onglets
        style = ttk.Style()
        style.configure("ConvFEA.TNotebook",
                         background=DARK_BG, borderwidth=0,
                         tabmargins=[0, 0, 0, 0])
        style.configure("ConvFEA.TNotebook.Tab",
                         background=PANEL_BG, foreground=TEXT_SEC,
                         font=FONT_LABEL, padding=[14, 6],
                         borderwidth=0)
        style.map("ConvFEA.TNotebook.Tab",
                   background=[("selected", CARD_BG)],
                   foreground=[("selected", ACCENT)])

        self._nb = ttk.Notebook(center, style="ConvFEA.TNotebook")
        self._nb.grid(row=0, column=0, sticky="nsew")

        # Onglet graphiques
        self._tab_charts = styled_frame(self._nb, fg_color=CARD_BG)
        self._nb.add(self._tab_charts, text="  Graphiques  ")
        self._build_charts_tab()

        # Onglet tableau
        self._tab_table = styled_frame(self._nb, fg_color=CARD_BG)
        self._nb.add(self._tab_table, text="  Tableau de résultats  ")
        self._build_table_tab()

        # Onglet historique ---
        self._tab_history = styled_frame(self._nb, fg_color=CARD_BG)
        self._nb.add(self._tab_history, text="  Historique  ")
        self._build_history_tab()

        # Console en bas
        console_frame = styled_frame(center, fg_color=PANEL_BG)
        console_frame.grid(row=1, column=0, sticky="ew", pady=(6, 0))
        label(console_frame, "Console d'exécution",
              font=FONT_TITLE, text_color=TEXT_SEC, fg_color=PANEL_BG).pack(
            anchor="w", padx=10, pady=(6, 2))
        self._console = tk.Text(console_frame, height=8,
                                 bg=DARK_BG, fg=ACCENT2, font=FONT_MONO,
                                 state="disabled", relief="flat", bd=0,
                                 insertbackground=ACCENT, wrap="word",
                                 padx=10, pady=6)
        self._console.pack(fill="both", expand=True, padx=6, pady=(0, 6))
        sb = ttk.Scrollbar(self._console, command=self._console.yview)
        self._console.configure(yscrollcommand=sb.set)

    def _build_charts_tab(self):
        frame = self._tab_charts
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        self._fig = Figure(figsize=(10, 6), facecolor="#1C2333")
        self._canvas = FigureCanvasTkAgg(self._fig, master=frame)
        self._canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        
        import tkinter as tk
        hidden_frame = tk.Frame(frame) 
        self._nav_toolbar = NavigationToolbar2Tk(self._canvas, hidden_frame)
        
        toolbar_frame = styled_frame(frame, fg_color="#1C2333", corner_radius=0)
        toolbar_frame.grid(row=1, column=0, sticky="ew")
        
        tools = [
            ("🏠", self._nav_toolbar.home),
            ("◀", self._nav_toolbar.back),
            ("▶", self._nav_toolbar.forward),
            ("✋", self._nav_toolbar.pan),
            ("🔍", self._nav_toolbar.zoom),
            ("💾", self._nav_toolbar.save_figure)
        ]
        
        for icon, cmd in tools:
            btn = ctk.CTkButton(
                toolbar_frame, 
                text=icon, 
                command=cmd, 
                width=36, height=30,
                fg_color="transparent",        # Fond invisible (se fond avec la barre)
                hover_color="#30363D",         # S'allume subtilement au survol
                text_color="#E6EDF3",          # Blanc
                font=ctk.CTkFont(size=18)      # Grande police pour les icônes
            )
            btn.pack(side="left", padx=2, pady=4)

        self._coord_label = ctk.CTkLabel(toolbar_frame, text="", text_color="#8B949E", 
                                         font=ctk.CTkFont("Segoe UI", 12))
        self._coord_label.pack(side="right", padx=14)

        def on_mouse_move(event):
            if event.inaxes:
                self._coord_label.configure(text=f"x = {event.xdata:.3f}   y = {event.ydata:.3f}")
            else:
                self._coord_label.configure(text="")
                
        self._canvas.mpl_connect("motion_notify_event", on_mouse_move)

        # Placeholder
        self._show_placeholder()

    def _build_table_tab(self):
        frame = self._tab_table
        frame.rowconfigure(1, weight=1)
        frame.columnconfigure(0, weight=1)

        # Filtre
        filter_row = styled_frame(frame, fg_color=CARD_BG)
        filter_row.grid(row=0, column=0, sticky="ew", padx=10, pady=(8, 4))
        label(filter_row, "Filtrer :", text_color=TEXT_SEC, fg_color=CARD_BG).pack(side="left")
        self.var_filter = tk.StringVar()
        self.var_filter.trace_add("write", lambda *_: self._filter_table())
        self._filter_entry = entry(filter_row, textvariable=self.var_filter, width=100)
        self._filter_entry.pack(side="left", padx=(6, 0))

        # Treeview
        style = ttk.Style()
        style.configure("ConvFEA.Treeview",
                         background=DARK_BG, foreground=TEXT_PRI,
                         rowheight=22, fieldbackground=DARK_BG,
                         borderwidth=0, font=FONT_MONO)
        style.configure("ConvFEA.Treeview.Heading",
                         background=PANEL_BG, foreground=ACCENT,
                         font=FONT_TITLE, relief="flat")
        style.map("ConvFEA.Treeview",
                   background=[("selected", CARD_BG)],
                   foreground=[("selected", ACCENT)])

        tv_frame = styled_frame(frame, fg_color=DARK_BG)
        tv_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))
        tv_frame.rowconfigure(0, weight=1)
        tv_frame.columnconfigure(0, weight=1)

        cols = COL_NAMES
        self._tv = ttk.Treeview(tv_frame, columns=cols, show="headings",
                                  style="ConvFEA.Treeview")
        for col in cols:
            short = col.replace(" (mm)", "").replace(" (MPa)", "").replace(" (mJ)", "")
            self._tv.heading(col, text=short,
                              command=lambda c=col: self._sort_table(c))
            self._tv.column(col, width=110, anchor="center", minwidth=60)
        self._tv.grid(row=0, column=0, sticky="nsew")

        vsb = ctk.CTkScrollbar(tv_frame, orientation="vertical", command=self._tv.yview)
        vsb.grid(row=0, column=1, sticky="ns")
        hsb = ctk.CTkScrollbar(tv_frame, orientation="horizontal", command=self._tv.xview)
        hsb.grid(row=1, column=0, sticky="ew")
        self._tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self._sort_col = None
        self._sort_reverse = False

    def _build_history_tab(self):
        self._tab_history.rowconfigure(0, weight=1)
        self._tab_history.columnconfigure(0, weight=1)
        
        self._history_scroll = ctk.CTkScrollableFrame(self._tab_history, fg_color=DARK_BG, corner_radius=8)
        self._history_scroll.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        
        self._refresh_history_ui()

    def _save_current_to_history(self):
        """Enregistre la simulation affichée dans la mémoire (History)."""
        if self.df is None:
            messagebox.showwarning("Vide", "Aucune donnée à enregistrer.")
            return
            
        name = self.current_loaded_name or self.var_simname.get()
        self.history[name] = self.df.copy()
        self._refresh_history_ui()
        self._log(f"✔ Simulation '{name}' enregistrée dans l'historique.")
        # Petit feedback visuel
        messagebox.showinfo("Succès", f"La simulation '{name}' a été sauvegardée dans l'historique.")

    def _refresh_history_ui(self):
        """Met à jour l'affichage de l'onglet Historique."""
        for w in self._history_scroll.winfo_children():
            w.destroy()
            
        if not self.history:
            label(self._history_scroll, "Aucune simulation enregistrée.", text_color=TEXT_DIM).pack(pady=20)
            return
            
        for name, df in self.history.items():
            card = styled_frame(self._history_scroll, fg_color=CARD_BG)
            card.pack(fill="x", padx=10, pady=6)
            
            info_frame = styled_frame(card, fg_color=CARD_BG)
            info_frame.pack(side="left", padx=10, pady=10)
            
            label(info_frame, name, font_size=12, text_color=ACCENT, bold=True).pack(anchor="w")
            label(info_frame, f"{len(df)} calculs de maillage", font_size=10, text_color=TEXT_SEC).pack(anchor="w")
            
            btn_frame = styled_frame(card, fg_color=CARD_BG)
            btn_frame.pack(side="right", padx=10, pady=10)
            
            accent_btn(btn_frame, "Charger", command=lambda n=name: self._load_from_history(n), width=80, height=28).pack(side="right")

    def _load_from_history(self, name):
        """Recharge une ancienne simulation dans l'interface principale."""
        if name in self.history:
            self.df = self.history[name].copy()
            self.current_loaded_name = name
            self.var_simname.set(name)
            
            self._update_charts()
            self._update_table()
            self._update_diagnostics()
            self._nb.select(0) # Ramène l'utilisateur sur l'onglet Graphiques
            self._log(f"✔ Simulation '{name}' rechargée depuis l'historique.")

    # ── Panneau droit (Conclusions) ──────────────────────────────────────────
    def _build_right_panel(self, parent):
        panel = styled_frame(parent, fg_color=PANEL_BG)
        panel.grid(row=0, column=2, sticky="nsew", padx=(6, 0))
        panel.configure(width=260)
        panel.pack_propagate(False)

        label(panel, "Diagnostic automatique",
              font=FONT_H1, text_color=TEXT_PRI, fg_color=PANEL_BG).pack(
            anchor="w", padx=14, pady=(14, 4))
        tk.Frame(panel, bg=BORDER, height=1).pack(fill="x", padx=10, pady=(0, 8))

        # Zone scrollable
        self._diag_inner = ctk.CTkScrollableFrame(panel, fg_color="#161B22", corner_radius=0)
        self._diag_inner.pack(fill="both", expand=True)

        self._show_diag_placeholder()

    # ─────────────────────────────────────────────────────────────────────────
    #  LOGIQUE DE L'INTERFACE
    # ─────────────────────────────────────────────────────────────────────────
    def _browse(self):
        path = filedialog.askopenfilename(
            title="Sélectionner le modèle Mechanical",
            filetypes=[("Fichiers Mechanical", "*.mechdat *.mechdb *.wbpj"),
                       ("Tous les fichiers", "*.*")]
        )
        if path:
            self.var_path.set(path)

    def get_mode(self):
        return {
            "Plage": "range",
            "Liste custom": "list"
        }[self.var_mode.get()]

    def _toggle_mesh_mode(self, value=None):
        if value is None:
            value = self.var_mode.get()

        self.var_mode.set(value)
        if self.get_mode() == "range":
            self._list_frame.pack_forget()
            self._range_frame.pack(fill="x")
        else:
            self._range_frame.pack_forget()
            self._list_frame.pack(fill="x")
        self._refresh_preview()

    def _get_sizes(self) -> list:
        if self.get_mode() == "range":
            mn  = float(self.var_min.get())
            mx  = float(self.var_max.get())
            stp = float(self.var_step.get())
            raw = []
            v = mx
            while v >= mn - 1e-9:
                raw.append(round(v, 6))
                v -= stp
            return sorted(set(raw), reverse=True)
        else:
            parts = [p.strip() for p in self.var_list.get().split(";") if p.strip()]
            return sorted([float(p) for p in parts], reverse=True)

    def _refresh_preview(self):
        try:
            sizes = self._get_sizes()
            txt = "Tailles (mm, ordre décroissant) :\n" + " → ".join(str(s) for s in sizes)
            txt += f"\n({len(sizes)} calculs)"
        except Exception as e:
            txt = f"Erreur : {e}"
        self._sizes_preview.configure(state="normal")
        self._sizes_preview.delete("1.0", "end")
        self._sizes_preview.insert("1.0", txt)
        self._sizes_preview.configure(state="disabled")

    def _log(self, msg: str):
        self._console.configure(state="normal")
        self._console.insert("end", msg + "\n")
        self._console.see("end")
        self._console.configure(state="disabled")

    def _set_progress(self, current: int, total: int):
        pct = current / total * 100 if total else 0
        self.var_prog.set(pct)
        self._prog_label.configure(text=f"Calcul {current}/{total} — {pct:.0f} %")

    # ── Lancement ────────────────────────────────────────────────────────────
    def _start_study(self):
        if self.running:
            return

        if self.df is not None:
            name = self.current_loaded_name or self.var_simname.get()
            self.history[name] = self.df.copy()
            self._refresh_history_ui()
            self._log(f"✔ Ancienne étude '{name}' auto-sauvegardée dans l'historique.")

        path = self.var_path.get().strip()
        if not path:
            messagebox.showwarning("Modèle manquant", "Veuillez sélectionner un fichier .mechdat.")
            return
        if not os.path.exists(path):
            messagebox.showwarning("Fichier introuvable", f"Le fichier n'existe pas :\n{path}")
            return
        try:
            sizes = self._get_sizes()
        except Exception as e:
            messagebox.showerror("Paramètres invalides", str(e))
            return
        if not sizes:
            messagebox.showwarning("Tailles vides", "Aucune taille de maillage définie.")
            return

        self.running = True
        self._run_btn.configure(state="disabled", text="⏳  Calcul en cours…")
        self.var_prog.set(0)
        self._prog_label.configure(text="Initialisation…")
        self._log("=" * 60)
        self._log(f"Simulation : {self.var_simname.get()}")
        self._log(f"Modèle     : {path}")
        self._log(f"Mode GUI   : {self.var_show_gui.get()}")
        self._log(f"Tailles    : {sizes}")
        self._log("=" * 60)

        # Extraction de la variable d'affichage de l'interface
        show_gui = self.var_show_gui.get()

        t = threading.Thread(
            target=run_study,
            args=(path, sizes, self.var_simname.get(), show_gui,
                  self._log_q, self._prog_q, self._res_q),
            daemon=True
        )
        t.start()

    # ── Polling des queues ────────────────────────────────────────────────────
    def _poll(self):
        # Messages log
        try:
            while True:
                msg = self._log_q.get_nowait()
                self._log(msg)
        except queue.Empty:
            pass

        # Progression
        try:
            while True:
                cur, tot = self._prog_q.get_nowait()
                self._set_progress(cur, tot)
        except queue.Empty:
            pass

        # Résultat final
        try:
            result = self._res_q.get_nowait()
            self._on_study_complete(result)
        except queue.Empty:
            pass

        self.after(150, self._poll)

    def _on_study_complete(self, result):
        self.running = False
        self._run_btn.configure(state="normal", text="▶  Lancer l'étude de convergence")

        if isinstance(result, Exception):
            self._log(f"✗ ERREUR : {result}")
            messagebox.showerror("Erreur de calcul", str(result))
            return

        self.df = result
        self.current_loaded_name = self.var_simname.get()
        self.var_prog.set(100)
        self._prog_label.configure(text=f"✔ Terminé — {len(self.df)} points")
        self._log(f"✔ Étude terminée : {len(self.df)} tailles calculées.")
        self._log("Mise à jour des graphiques et du tableau…")

        self._update_charts()
        self._update_table()
        self._update_diagnostics()
        self._nb.select(0)

    # ─────────────────────────────────────────────────────────────────────────
    #  VISUALISATION
    # ─────────────────────────────────────────────────────────────────────────
    def _show_placeholder(self):
        self._fig.clear()
        ax = self._fig.add_subplot(111)
        ax.set_facecolor(CARD_BG)
        ax.text(0.5, 0.5,
                "Les graphiques s'afficheront ici\naprès l'exécution de l'étude.",
                ha="center", va="center", color=TEXT_DIM,
                fontsize=12, transform=ax.transAxes,
                fontfamily="Segoe UI")
        ax.axis("off")
        self._fig.patch.set_facecolor(CARD_BG)
        self._canvas.draw()

    def _update_charts(self):
        if self.df is None:
            return

        # 1. Trier les données par taille croissante (strictement requis pour le lissage)
        df = self.df.sort_values(by="Nombre d'éléments")
        x = df["Nombre d'éléments"].values

        self._fig.clear()
        self._fig.patch.set_facecolor(CARD_BG)

        n = len(PLOT_CONFIGS)
        ncols = 4
        nrows = (n + ncols - 1) // ncols
        gs = self._fig.add_gridspec(nrows, ncols,
                                     hspace=0.6, wspace=0.4,
                                     left=0.06, right=0.97,
                                     top=0.94, bottom=0.08)

        # Liste pour stocker les marqueurs et y attacher les infobulles interactives
        scatter_points = []

        for i, (title, ycols, ylabel) in enumerate(PLOT_CONFIGS):
            r, c = divmod(i, ncols)
            ax = self._fig.add_subplot(gs[r, c])
            ax.set_facecolor(DARK_BG)

            for j, col in enumerate(ycols):
                if col in df.columns:
                    color = CURVE_COLORS[i % len(CURVE_COLORS)]
                    style = "-" if j == 0 else "--"
                    y = df[col].values

                    # Lissage des courbes (Interpolation monotone façon Excel/Plotly)
                    if len(x) >= 4:
                        try:
                            # PchipInterpolator évite les "overshoots" (rebonds mathématiques) 
                            # non physiques, idéal pour des courbes de convergence FEA.
                            x_smooth = np.linspace(x.min(), x.max(), 300)
                            pchip = PchipInterpolator(x, y)
                            y_smooth = pchip(x_smooth)

                            # Tracé de la courbe lissée
                            ax.plot(x_smooth, y_smooth, linewidth=1.8, linestyle=style, color=color, alpha=0.8)
                        except Exception:
                            # Sécurité : ligne droite si l'interpolation échoue
                            ax.plot(x, y, linewidth=1.8, linestyle=style, color=color, alpha=0.8)
                    else:
                        ax.plot(x, y, linewidth=1.8, linestyle=style, color=color, alpha=0.8)

                    # Marqueurs servant de points d'accroche pour l'interactivité
                    scat, = ax.plot(x, y, marker="o", markersize=5, linewidth=0,
                                    color=color, label=col.split("(")[0].strip())
                    scatter_points.append(scat)

            ax.set_title(title, color=TEXT_PRI, fontsize=7.5, pad=4, fontfamily="Segoe UI")
            ax.set_xlabel("Nombre d'éléments", color=TEXT_DIM, fontsize=6.5)
            ax.set_ylabel(ylabel, color=TEXT_DIM, fontsize=6.5)
            ax.tick_params(colors=TEXT_DIM, labelsize=6)
            
            for spine in ax.spines.values():
                spine.set_edgecolor(BORDER)
            ax.grid(True, color=BORDER, linewidth=0.5, alpha=0.7)

            if len(ycols) > 1:
                ax.legend(fontsize=5.5, facecolor=CARD_BG, edgecolor=BORDER, labelcolor=TEXT_SEC)

        # Affichage dynamique des valeurs (Hover interactif)
        try:
            cursor = mplcursors.cursor(scatter_points, hover=True)
            @cursor.connect("add")
            def on_add(sel):
                # Formatage de l'infobulle pour un style moderne
                sel.annotation.set_text(f"Éléments : {int(sel.target[0])}\nValeur : {sel.target[1]:.4f}")
                sel.annotation.get_bbox_patch().set(
                    fc=PANEL_BG, ec=ACCENT, alpha=0.9, boxstyle="round,pad=0.4"
                )
                sel.annotation.set_color(TEXT_PRI)
                sel.annotation.set_fontsize(8)

            def on_leave(event):
                # Si une bulle est active, on l'efface
                for sel in cursor.selections:
                    cursor.remove_selection(sel)
                self._canvas.draw_idle() # Demande à Matplotlib de rafraîchir l'écran
                
            # On écoute la sortie de la souris hors de la figure et des axes
            self._canvas.mpl_connect("figure_leave_event", on_leave)
            self._canvas.mpl_connect("axes_leave_event", on_leave)

        except Exception:
            pass # Ignore en cas d'absence de focus GUI

        self._canvas.draw()

    # ─────────────────────────────────────────────────────────────────────────
    #  TABLEAU
    # ─────────────────────────────────────────────────────────────────────────
    def _update_table(self):
        self._populate_treeview(self.df)

    def _populate_treeview(self, df):
        for row in self._tv.get_children():
            self._tv.delete(row)
        if df is None:
            return
        for _, r in df.iterrows():
            vals = []
            for v in r:
                try:
                    vals.append(f"{float(v):.4f}")
                except Exception:
                    vals.append(str(v))
            self._tv.insert("", "end", values=vals)

    def _filter_table(self):
        if self.df is None:
            return
        filt = self.var_filter.get().lower()
        if not filt:
            self._populate_treeview(self.df)
            return
        filtered = self.df[
            self.df.apply(
                lambda row: any(filt in str(v).lower() for v in row), axis=1
            )
        ]
        self._populate_treeview(filtered)

    def _sort_table(self, col):
        if self.df is None:
            return
        if self._sort_col == col:
            self._sort_reverse = not self._sort_reverse
        else:
            self._sort_col = col
            self._sort_reverse = False
        df_sorted = self.df.sort_values(col, ascending=not self._sort_reverse)
        self._populate_treeview(df_sorted)

    # ─────────────────────────────────────────────────────────────────────────
    #  DIAGNOSTIC
    # ─────────────────────────────────────────────────────────────────────────
    def _show_diag_placeholder(self):
        for w in self._diag_inner.winfo_children():
            w.destroy()
        label(self._diag_inner,
              "Le diagnostic s'affichera\naprès l'étude de convergence.",
              text_color=TEXT_DIM, fg_color=PANEL_BG, font=FONT_LABEL, justify="left").pack(
            padx=10, pady=10, anchor="w")

    def _update_diagnostics(self):
        for w in self._diag_inner.winfo_children():
            w.destroy()
        if self.df is None:
            self._show_diag_placeholder()
            return

        msgs = build_diagnostics(self.df)
        for icon, color, text in msgs:
            card = styled_frame(self._diag_inner, fg_color=CARD_BG)
            card.pack(fill="x", padx=8, pady=3)
            # Bande colorée
            tk.Frame(card, bg=color, width=3).pack(side="left", fill="y")
            inner = styled_frame(card, fg_color=CARD_BG)
            inner.pack(side="left", fill="both", expand=True, padx=(8, 8), pady=6)
            label(inner, icon, font=("Segoe UI", 12), text_color=color, fg_color=CARD_BG).pack(
                anchor="w")
            wrapped = textwrap.fill(text, width=30)
            label(inner, wrapped, font=("Segoe UI", 8), text_color=TEXT_SEC,
                  fg_color=CARD_BG, justify="left").pack(anchor="w")

        # Erreurs relatives
        sep = styled_frame(self._diag_inner, fg_color=PANEL_BG)
        sep.pack(fill="x", padx=8, pady=(8, 0))
        tk.Frame(sep, bg=BORDER, height=1).pack(fill="x")
        label(sep, "Erreurs relatives (derniers raffinements)",
              font=("Segoe UI", 8, "bold"), text_color=TEXT_SEC, fg_color=PANEL_BG).pack(
            anchor="w", padx=0, pady=(4, 2))

        cols_check = [
            "Contrainte max Von Mises (MPa)",
            "Contrainte moy Von Mises (MPa)",
            "Énergie de déformation (mJ)",
            "Déplacement maximal (mm)",
        ]
        for col in cols_check:
            if col in self.df.columns and len(self.df) >= 2:
                vals = self.df[col].values
                ref  = vals[-1]
                if ref != 0:
                    err = abs(vals[-1] - vals[-2]) / abs(ref) * 100
                    short = col.split("(")[0].strip()
                    color = ACCENT2 if err < 5 else (WARN if err > 15 else "#FFA657")
                    row_f = styled_frame(self._diag_inner, fg_color=PANEL_BG)
                    row_f.pack(fill="x", padx=8)
                    label(row_f, f"{short[:22]}:", text_color=TEXT_DIM, fg_color=PANEL_BG,
                           font=("Segoe UI", 7.5)).pack(side="left")
                    label(row_f, f"{err:.1f} %", text_color=color, fg_color=PANEL_BG,
                           font=("Segoe UI Semibold", 7.5)).pack(side="right")

    # ─────────────────────────────────────────────────────────────────────────
    #  EXPORT
    # ─────────────────────────────────────────────────────────────────────────
    def _check_data(self) -> bool:
        if self.df is None:
            messagebox.showwarning("Pas de données", "Lancez d'abord une étude de convergence.")
            return False
        return True

    def _export_csv(self):
        if not self._check_data():
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            initialfile=f"{self.var_simname.get()}.csv",
            filetypes=[("CSV", "*.csv")]
        )
        if path:
            self.df.to_csv(path, index=False, sep=";", decimal=",")
            self._log(f"✔ CSV exporté : {path}")
            messagebox.showinfo("Export CSV", f"Fichier enregistré :\n{path}")

    def _export_excel(self):
        if not self._check_data():
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=f"{self.var_simname.get()}.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = self.var_simname.get()[:31]

        # En-têtes
        for ci, col in enumerate(COL_NAMES, 1):
            ws.cell(row=1, column=ci, value=col)

        # Données
        for ri, row in self.df.iterrows():
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri + 2, column=ci, value=val)

        wb.save(path)
        self._log(f"✔ Excel exporté : {path}")
        messagebox.showinfo("Export Excel", f"Fichier enregistré :\n{path}")

    def _export_png(self):
        if not self._check_data():
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".png",
            initialfile=f"{self.var_simname.get()}_graphiques.png",
            filetypes=[("PNG", "*.png"), ("SVG", "*.svg")]
        )
        if path:
            self._fig.savefig(path, dpi=200, bbox_inches="tight",
                              facecolor=CARD_BG)
            self._log(f"✔ Graphique exporté : {path}")
            messagebox.showinfo("Export graphique", f"Fichier enregistré :\n{path}")

    def _export_pdf(self):
        if not self._check_data():
            return
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                             Table, TableStyle, Image as RLImage,
                                             HRFlowable)
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            import io
        except ImportError:
            messagebox.showerror("Dépendance manquante",
                                  "Installez reportlab :\n  pip install reportlab")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            initialfile=f"{self.var_simname.get()}_rapport.pdf",
            filetypes=[("PDF", "*.pdf")]
        )
        if not path:
            return

        doc = SimpleDocTemplate(path, pagesize=A4,
                                 leftMargin=2*cm, rightMargin=2*cm,
                                 topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        elements = []

        # Titre
        title_style = ParagraphStyle("Title2", parent=styles["Title"],
                                      fontSize=18, textColor=colors.HexColor("#1a73e8"))
        elements.append(Paragraph(f"Rapport de convergence FEA", title_style))
        elements.append(Paragraph(f"Simulation : {self.var_simname.get()}",
                                   styles["Heading2"]))
        elements.append(Paragraph(f"Date : {time.strftime('%Y-%m-%d %H:%M')}",
                                   styles["Normal"]))
        elements.append(HRFlowable(width="100%", thickness=1,
                                    color=colors.HexColor("#30363D")))
        elements.append(Spacer(1, 0.3*cm))

        # Tableau de résultats
        elements.append(Paragraph("Résultats par taille de maillage", styles["Heading2"]))
        short_cols = [c.replace(" (mm)", "").replace(" (MPa)", "").replace(" (mJ)", "")
                      for c in COL_NAMES]
        data = [short_cols]
        for _, row in self.df.iterrows():
            data.append([f"{float(v):.3f}" for v in row])

        tbl = Table(data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a73e8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 6),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.HexColor("#f8f9fa"), colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dee2e6")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(tbl)
        elements.append(Spacer(1, 0.5*cm))

        # Graphique
        elements.append(Paragraph("Courbes de convergence", styles["Heading2"]))
        img_buf = io.BytesIO()
        self._fig.savefig(img_buf, format="png", dpi=150, bbox_inches="tight",
                          facecolor=CARD_BG)
        img_buf.seek(0)
        elements.append(RLImage(img_buf, width=17*cm,
                                 height=17*cm * self._fig.get_figheight() / self._fig.get_figwidth()))
        elements.append(Spacer(1, 0.5*cm))

        # Diagnostic
        elements.append(Paragraph("Diagnostic automatique", styles["Heading2"]))
        msgs = build_diagnostics(self.df)
        for icon, _, text in msgs:
            elements.append(Paragraph(f"{icon} {text}", styles["Normal"]))
            elements.append(Spacer(1, 0.15*cm))

        doc.build(elements)
        self._log(f"✔ PDF exporté : {path}")
        messagebox.showinfo("Export PDF", f"Rapport enregistré :\n{path}")


# ─────────────────────────────────────────────────────────────────────────────
#  POINT D'ENTRÉE
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = ConvFEA()
    app.mainloop()
